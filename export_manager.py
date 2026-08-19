"""
Module d'exportation comptable pour L'ADRESSE B.
Toutes les requêtes SQL utilisent des prepared statements.
"""
import csv
import datetime
import os
import json
from decimal import Decimal
from database_manager import get_connection

EXPORT_DIR = "Exports_L_ADRESSE_B"

def _ensure_dir():
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)

def export_comptable_belge():
    """
    Exporte un CSV normé comptabilité belge.
    Colonnes : Date, Numéro Ticket, Code Article, Nom Article,
               Montant TVAC, Base HTVA, Montant TVA, Moyen de Paiement.
    Retourne le chemin du fichier créé ou lève une exception.
    """
    conn = get_connection()
    c = conn.cursor()
    # Jointure Tickets × Ventes_Details × Produits via Stocks
    c.execute("""
        SELECT
            t.date_heure,
            t.numero_ticket,
            p.code_barre,
            p.nom,
            vd.prix_unitaire_tvac,
            p.taux_tva,
            t.methode_paiement
        FROM Tickets t
        JOIN Ventes_Details vd ON vd.id_ticket = t.id
        JOIN Stocks s          ON s.id = vd.id_stock
        JOIN Produits p        ON p.id = s.id_produit
        ORDER BY t.date_heure DESC
    """)
    rows = c.fetchall()
    conn.close()

    records = []
    for date_heure, ticket, code, nom, tvac, taux, methode in rows:
        tvac_d = Decimal(str(tvac))
        taux_d = Decimal(str(taux))
        htva   = (tvac_d / (Decimal("1") + taux_d)).quantize(Decimal("0.0001"))
        tva    = (tvac_d - htva).quantize(Decimal("0.0001"))
        records.append({
            "Date":              date_heure,
            "Numéro Ticket":     ticket,
            "Code Article":      code,
            "Nom Article":       nom,
            "Montant TVAC (€)":  str(tvac_d).replace(".", ","),
            "Base HTVA (€)":     str(htva).replace(".", ","),
            "Montant TVA (€)":   str(tva).replace(".", ","),
            "Taux TVA":          f"{float(taux_d)*100:.0f}%",
            "Moyen de Paiement": methode,
        })

    _ensure_dir()
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"export_comptable_{ts}.csv")
    fieldnames = [
        "Date", "Numéro Ticket", "Code Article", "Nom Article",
        "Montant TVAC (€)", "Base HTVA (€)", "Montant TVA (€)", "Taux TVA", "Moyen de Paiement"
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(records)
    return path


def export_winbooks_csv(mois=None, annee=None):
    """
    Exporte un CSV formaté pour les logiciels comptables belges (ex: WinBooks, Exact).
    Génère des écritures en partie double (Débit/Crédit) avec des comptes généraux par défaut.
    """
    conn = get_connection()
    c = conn.cursor()
    
    query = """
        SELECT
            numero_ticket,
            date_heure,
            total_tvac,
            total_htva,
            total_tva,
            methode_paiement
        FROM Tickets
    """
    params = []
    if mois and annee:
        query += " WHERE strftime('%m', date_heure) = ? AND strftime('%Y', date_heure) = ?"
        params.extend([f"{int(mois):02d}", str(annee)])
        
    query += " ORDER BY date_heure DESC"
    
    c.execute(query, params)
    rows = c.fetchall()
    conn.close()

    records = []
    for num, date_heure, tvac, htva, tva, methode in rows:
        date_str = str(date_heure)[:10]
        periode = date_str[:7].replace("-", "")
        date_format = date_str[8:10] + "/" + date_str[5:7] + "/" + date_str[:4]
        
        # 1. Ligne Débit (Paiement)
        # 580000 pour virement/carte, 570000 pour caisse espèces
        compte_debit = "570000" if methode == "Espèces" else "580000"
        records.append({
            "Code Journal": "VEN",
            "Période": periode,
            "Date": date_format,
            "Document": num,
            "Compte": compte_debit,
            "Libellé": f"Recette {methode}",
            "Sens": "D",
            "Montant": str(Decimal(str(tvac))).replace(".", ","),
            "Code TVA": ""
        })
        
        # 2. Ligne Crédit (Vente HTVA)
        records.append({
            "Code Journal": "VEN",
            "Période": periode,
            "Date": date_format,
            "Document": num,
            "Compte": "700000",
            "Libellé": "Vente Marchandises",
            "Sens": "C",
            "Montant": str(Decimal(str(htva))).replace(".", ","),
            "Code TVA": "21"
        })
        
        # 3. Ligne Crédit (TVA Due)
        if float(tva) > 0:
            records.append({
                "Code Journal": "VEN",
                "Période": periode,
                "Date": date_format,
                "Document": num,
                "Compte": "451000",
                "Libellé": "TVA à payer",
                "Sens": "C",
                "Montant": str(Decimal(str(tva))).replace(".", ","),
                "Code TVA": ""
            })

    _ensure_dir()
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"export_winbooks_{ts}.csv")
    fieldnames = [
        "Code Journal", "Période", "Date", "Document", "Compte", "Libellé", "Sens", "Montant", "Code TVA"
    ]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(records)
    return path



def export_synthese_gerant(comptage_details=None):
    """
    Génère un fichier .xlsx de clôture journalière.
    Feuilles :
      1. Synthèse  — CA, répartition par méthode de paiement
      2. Top 5     — 5 articles les plus vendus du jour
      3. Détail    — Toutes les ventes du jour
    Retourne le chemin du fichier créé ou lève une exception.
    """
    today = datetime.date.today().isoformat()
    conn  = get_connection()
    c     = conn.cursor()

    # ── CA et répartition par méthode
    c.execute("""
        SELECT methode_paiement,
               COUNT(*)              AS nb_tickets,
               SUM(total_tvac)       AS total_tvac,
               SUM(total_htva)       AS total_htva,
               SUM(total_tva)        AS total_tva
        FROM Tickets
        WHERE date(date_heure) = ?
        GROUP BY methode_paiement
    """, (today,))
    repartition = c.fetchall()

    # ── Top 5 articles du jour
    c.execute("""
        SELECT p.nom, p.categorie, SUM(vd.quantite) AS qte_vendue,
               SUM(vd.prix_unitaire_tvac * vd.quantite) AS ca_genere
        FROM Ventes_Details vd
        JOIN Stocks s   ON s.id = vd.id_stock
        JOIN Produits p ON p.id = s.id_produit
        JOIN Tickets t  ON t.id = vd.id_ticket
        WHERE date(t.date_heure) = ?
        GROUP BY p.id
        ORDER BY qte_vendue DESC
        LIMIT 5
    """, (today,))
    top5 = c.fetchall()

    # ── Détail complet du jour
    c.execute("""
        SELECT t.date_heure, t.numero_ticket, p.nom, p.code_barre,
               vd.prix_unitaire_tvac, t.methode_paiement
        FROM Tickets t
        JOIN Ventes_Details vd ON vd.id_ticket = t.id
        JOIN Stocks s          ON s.id = vd.id_stock
        JOIN Produits p        ON p.id = s.id_produit
        WHERE date(t.date_heure) = ?
        ORDER BY t.date_heure
    """, (today,))
    detail = c.fetchall()
    conn.close()

    _ensure_dir()
    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(EXPORT_DIR, f"cloture_journaliere_{ts}.xlsx")

    import openpyxl
    wb = openpyxl.Workbook()

    # Feuille 1 — Synthèse
    ws1 = wb.active
    ws1.title = "Synthèse"
    ws1.append(["CLÔTURE DE CAISSE — L'ADRESSE B"])
    ws1.append([f"Date : {datetime.date.today().strftime('%d/%m/%Y')}"])
    ws1.append([])
    headers1 = ["Méthode", "Nb Tickets", "Total TVAC (€)", "Total HTVA (€)", "Total TVA (€)"]
    ws1.append(headers1)

    tot_nb = 0
    tot_tvac = Decimal("0.00")
    tot_htva = Decimal("0.00")
    tot_tva = Decimal("0.00")
    for row in repartition:
        meth = row[0] or "Autre"
        nb = row[1] or 0
        tvac = Decimal(str(row[2] or 0))
        htva = Decimal(str(row[3] or 0))
        tva = Decimal(str(row[4] or 0))
        tot_nb += nb
        tot_tvac += tvac
        tot_htva += htva
        tot_tva += tva
        ws1.append([meth, nb, float(tvac), float(htva), float(tva)])
    ws1.append(["TOTAL", tot_nb, float(tot_tvac), float(tot_htva), float(tot_tva)])
    _style_header(ws1, row=4, ncols=5)

    # Feuille 2 — Top 5
    ws2 = wb.create_sheet(title="Top 5 Articles")
    ws2.append(["Article", "Catégorie", "Qté Vendue", "CA Généré (€)"])
    for row in top5:
        ws2.append([row[0], row[1], row[2], float(Decimal(str(row[3] or 0)))])
    _style_header(ws2, row=1, ncols=4)

    # Feuille 3 — Détail
    ws3 = wb.create_sheet(title="Détail Ventes")
    ws3.append(["Date/Heure", "Ticket", "Article", "Code Barre", "Prix TVAC (€)", "Paiement"])
    for row in detail:
        ws3.append([row[0], row[1], row[2], row[3], float(Decimal(str(row[4] or 0))), row[5]])
    _style_header(ws3, row=1, ncols=6)

    # Feuille 4 — Comptage
    if comptage_details:
        records_comptage = []
        for val_str, qte in comptage_details.items():
            if qte > 0:
                val_num = float(val_str)
                records_comptage.append((val_num, f"{val_num:.2f} €" if val_num < 5 else f"{int(val_num)} €", qte, qte * val_num))
        if records_comptage:
            records_comptage.sort(key=lambda x: x[0], reverse=True)
            ws4 = wb.create_sheet(title="Comptage Caisse")
            ws4.append(["Dénomination", "Quantité", "Sous-Total (€)"])
            tot_c = 0.0
            for r in records_comptage:
                ws4.append([r[1], r[2], r[3]])
                tot_c += r[3]
            ws4.append(["TOTAL COMPTÉ", "", tot_c])
            _style_header(ws4, row=1, ncols=3)

    wb.save(path)
    return path


def _style_header(ws, row, ncols):
    """Applique un style minimal à la ligne d'en-tête."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor="1A1A1A")
        font = Font(bold=True, color="D4AF37")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font  = fill and font
            cell.fill  = fill
            cell.alignment = Alignment(horizontal="center")
        # Largeur auto
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    except Exception:
        pass  # Style non critique


def sauvegarder_rapport_z_journalier(comptage_details=None):
    """
    Génère et stocke le rapport Z consolidé de la journée en cours
    dans la table immuable Rapports_Z au format JSON.
    """
    conn = get_connection()
    c = conn.cursor()
    
    # Date du jour (ou de la session courante)
    today = datetime.date.today().isoformat()
    
    # Informations de session
    c.execute("SELECT fond_caisse_matin, montant_compté_soir, montant_theorique_soir, ecart_caisse FROM Sessions_Caisse ORDER BY id DESC LIMIT 1")
    session_row = c.fetchone()
    fond_initial = float(session_row[0]) if session_row and session_row[0] is not None else 0.0
    solde_reel = float(session_row[1]) if session_row and session_row[1] is not None else 0.0
    solde_theo = float(session_row[2]) if session_row and session_row[2] is not None else 0.0
    ecart = float(session_row[3]) if session_row and session_row[3] is not None else 0.0
    
    # Sorties manuelles
    c.execute("SELECT COALESCE(SUM(montant), 0) FROM Depenses_Caisse WHERE date(date_heure) = ?", (today,))
    sorties_manuelles = float(c.fetchone()[0])
    
    # Récupération des tickets
    c.execute("SELECT id, numero_ticket, total_tvac, total_htva, total_tva, methode_paiement FROM Tickets WHERE date(date_heure) = ? ORDER BY id ASC", (today,))
    tickets = c.fetchall()
    
    premier_ticket = tickets[0][1] if tickets else ""
    dernier_ticket = tickets[-1][1] if tickets else ""
    
    ca_ttc_total = Decimal("0.00")
    ventilation_tva = {
        "0.21": {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")},
        "0.06": {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")}
    }
    
    repartition_paiements = {
        "Bancontact/Visa": Decimal("0.00"),
        "Espèces": Decimal("0.00"),
        "Cartes Cadeaux / Autres": Decimal("0.00"),
        "Remboursements": Decimal("0.00")
    }
    
    for t_id, num, tvac_val, htva_val, tva_val, methode in tickets:
        tvac_d = Decimal(str(tvac_val)) if tvac_val is not None else Decimal("0.00")
        ca_ttc_total += tvac_d
        
        # Mapping méthode de paiement
        m_str = str(methode)
        if "REMB" in m_str or tvac_d < 0:
            repartition_paiements["Remboursements"] += tvac_d
        elif "Bancontact" in m_str or "Visa" in m_str:
            repartition_paiements["Bancontact/Visa"] += tvac_d
        elif "Espèces" in m_str:
            repartition_paiements["Espèces"] += tvac_d
        else:
            repartition_paiements["Cartes Cadeaux / Autres"] += tvac_d
            
        # Détails des lignes pour ventiler la TVA avec précision
        c.execute("""
            SELECT vd.prix_unitaire_tvac, vd.quantite, p.taux_tva
            FROM Ventes_Details vd
            JOIN Stocks s ON s.id = vd.id_stock
            JOIN Produits p ON p.id = s.id_produit
            WHERE vd.id_ticket = ?
        """, (t_id,))
        lignes = c.fetchall()
        
        if lignes:
            for p_tvac, qte, taux in lignes:
                t_taux = str(taux) if taux is not None else "0.21"
                if t_taux not in ventilation_tva:
                    ventilation_tva[t_taux] = {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")}
                
                l_ttc = Decimal(str(p_tvac)) * Decimal(str(qte))
                l_htva = (l_ttc / (Decimal("1") + Decimal(t_taux))).quantize(Decimal("0.01"))
                l_tva = l_ttc - l_htva
                
                ventilation_tva[t_taux]["ttc"] += l_ttc
                ventilation_tva[t_taux]["base_htva"] += l_htva
                ventilation_tva[t_taux]["montant_tva"] += l_tva
        else:
            # Ticket sans lignes (ex. Remboursement direct)
            # On utilise le taux par défaut de 21%
            t_taux = "0.21"
            l_ttc = tvac_d
            l_htva = (l_ttc / Decimal("1.21")).quantize(Decimal("0.01"))
            l_tva = l_ttc - l_htva
            
            ventilation_tva[t_taux]["ttc"] += l_ttc
            ventilation_tva[t_taux]["base_htva"] += l_htva
            ventilation_tva[t_taux]["montant_tva"] += l_tva
            
    # Construction de l'objet JSON Z
    objet_z = {
        "date": today,
        "sequence": {"premier_ticket": premier_ticket, "dernier_ticket": dernier_ticket},
        "financier": {
            "ca_ttc": float(ca_ttc_total),
            "tva": {k: {"base_htva": float(v["base_htva"]), "montant_tva": float(v["montant_tva"]), "ttc": float(v["ttc"])} for k, v in ventilation_tva.items()}
        },
        "paiements": {k: float(v) for k, v in repartition_paiements.items()},
        "mouvements_caisse": {
            "fond_initial": fond_initial,
            "sorties_manuelles": sorties_manuelles,
            "solde_theorique": solde_theo,
            "solde_reel": solde_reel,
            "ecart": ecart
        },
        "statut": {
            "caisse_juste": (abs(ecart) < 0.01)
        },
        "comptage_details": comptage_details or {}
    }
    
    donnees_json = json.dumps(objet_z, ensure_ascii=False)
    
    from database_manager import signer_rapport_z
    sig_z, hash_z = signer_rapport_z(c, today, donnees_json)
    
    # Ajoute les colonnes de signature dans la table
    c.execute("INSERT OR REPLACE INTO Rapports_Z (date, donnees_json, signature, hash_precedent) VALUES (?, ?, ?, ?)", (today, donnees_json, sig_z, hash_z))
    conn.commit()
    conn.close()
    return objet_z



def calculer_rapport_z_virtuel(date_str, c):
    c.execute("SELECT fond_caisse_matin FROM Sessions_Caisse WHERE date(date_ouverture) <= ? ORDER BY id DESC LIMIT 1", (date_str,))
    session_row = c.fetchone()
    fond_initial = float(session_row[0]) if session_row and session_row[0] is not None else 0.0
    
    c.execute("SELECT COALESCE(SUM(montant), 0) FROM Depenses_Caisse WHERE date(date_heure) = ?", (date_str,))
    sorties_manuelles = float(c.fetchone()[0])
    
    c.execute("SELECT id, numero_ticket, total_tvac, total_htva, total_tva, methode_paiement FROM Tickets WHERE date(date_heure) = ? ORDER BY id ASC", (date_str,))
    tickets = c.fetchall()
    
    if not tickets and sorties_manuelles == 0:
        return None
        
    premier_ticket = tickets[0][1] if tickets else ""
    dernier_ticket = tickets[-1][1] if tickets else ""
    
    from decimal import Decimal
    ca_ttc_total = Decimal("0.00")
    ventilation_tva = {
        "0.21": {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")},
        "0.06": {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")}
    }
    
    repartition_paiements = {
        "Bancontact/Visa": Decimal("0.00"),
        "Espèces": Decimal("0.00"),
        "Cartes Cadeaux / Autres": Decimal("0.00"),
        "Remboursements": Decimal("0.00")
    }
    
    for t_id, num, tvac_val, htva_val, tva_val, methode in tickets:
        tvac_d = Decimal(str(tvac_val)) if tvac_val is not None else Decimal("0.00")
        ca_ttc_total += tvac_d
        
        m_str = str(methode)
        if "REMB" in m_str or tvac_d < 0:
            repartition_paiements["Remboursements"] += tvac_d
        elif "Bancontact" in m_str or "Visa" in m_str or "QR" in m_str:
            repartition_paiements["Bancontact/Visa"] += tvac_d
        elif "Espèces" in m_str:
            repartition_paiements["Espèces"] += tvac_d
        else:
            repartition_paiements["Cartes Cadeaux / Autres"] += tvac_d
            
        c.execute("""
            SELECT vd.prix_unitaire_tvac, vd.quantite, p.taux_tva
            FROM Ventes_Details vd
            JOIN Stocks s ON s.id = vd.id_stock
            JOIN Produits p ON p.id = s.id_produit
            WHERE vd.id_ticket = ?
        """, (t_id,))
        lignes = c.fetchall()
        
        if lignes:
            for p_tvac, qte, taux in lignes:
                t_taux = str(taux) if taux is not None else "0.21"
                if t_taux not in ventilation_tva:
                    ventilation_tva[t_taux] = {"base_htva": Decimal("0.00"), "montant_tva": Decimal("0.00"), "ttc": Decimal("0.00")}
                
                l_ttc = Decimal(str(p_tvac)) * Decimal(str(qte))
                l_htva = (l_ttc / (Decimal("1") + Decimal(t_taux))).quantize(Decimal("0.01"))
                l_tva = l_ttc - l_htva
                
                ventilation_tva[t_taux]["ttc"] += l_ttc
                ventilation_tva[t_taux]["base_htva"] += l_htva
                ventilation_tva[t_taux]["montant_tva"] += l_tva
        else:
            t_taux = "0.21"
            l_ttc = tvac_d
            l_htva = (l_ttc / Decimal("1.21")).quantize(Decimal("0.01"))
            l_tva = l_ttc - l_htva
            
            ventilation_tva[t_taux]["ttc"] += l_ttc
            ventilation_tva[t_taux]["base_htva"] += l_htva
            ventilation_tva[t_taux]["montant_tva"] += l_tva
            
    return {
        "date": date_str,
        "sequence": {"premier_ticket": premier_ticket, "dernier_ticket": dernier_ticket},
        "financier": {
            "ca_ttc": float(ca_ttc_total),
            "tva": {k: {"base_htva": float(v["base_htva"]), "montant_tva": float(v["montant_tva"]), "ttc": float(v["ttc"])} for k, v in ventilation_tva.items()}
        },
        "paiements": {k: float(v) for k, v in repartition_paiements.items()},
        "mouvements_caisse": {
            "fond_initial": fond_initial,
            "sorties_manuelles": sorties_manuelles,
            "solde_theorique": float(Decimal(str(fond_initial)) + repartition_paiements["Espèces"] - Decimal(str(sorties_manuelles))),
            "solde_reel": float(Decimal(str(fond_initial)) + repartition_paiements["Espèces"] - Decimal(str(sorties_manuelles))),
            "ecart": 0.0
        },
        "statut": {"caisse_juste": True},
        "comptage_details": {}
    }


def export_comptable_mensuel(mois, annee, format_type="excel"):
    """
    Génère l'export mensuel agrégé pour l'expert-comptable
    en consommant les rapports Z archivés et en y ajoutant les jours en temps réel (virtuels)
    non encore clôturés.
    format_type: 'excel', 'csv' ou 'json'
    """
    conn = get_connection()
    c = conn.cursor()
    # Format mois sur 2 chiffres
    mois_str = f"{int(mois):02d}"
    annee_str = str(annee)
    
    c.execute("SELECT date, donnees_json FROM Rapports_Z WHERE strftime('%m', date) = ? AND strftime('%Y', date) = ? ORDER BY date ASC", (mois_str, annee_str))
    lignes_z = c.fetchall()
    
    # Check for unclosed tickets and expenses
    c.execute("SELECT DISTINCT date(date_heure) FROM Tickets WHERE strftime('%m', date_heure) = ? AND strftime('%Y', date_heure) = ?", (mois_str, annee_str))
    jours_tickets = [r[0] for r in c.fetchall()]
    
    c.execute("SELECT DISTINCT date(date_heure) FROM Depenses_Caisse WHERE strftime('%m', date_heure) = ? AND strftime('%Y', date_heure) = ?", (mois_str, annee_str))
    jours_depenses = [r[0] for r in c.fetchall()]
    
    jours_actifs = set(jours_tickets + jours_depenses)
    jours_avec_z = {date_z for date_z, _ in lignes_z}
    
    import json
    for jour in jours_actifs:
        if jour not in jours_avec_z:
            virtual_z = calculer_rapport_z_virtuel(jour, c)
            if virtual_z:
                lignes_z.append((jour, json.dumps(virtual_z)))
    
    lignes_z.sort(key=lambda x: x[0])
    conn.close()
    
    if not lignes_z:
        # Permettre l'exportation d'un rapport comptable à zéro même si aucun ticket n'a été émis
        virtual_z = {
            "financier": {"ca_ttc": 0.0, "tva": {}},
            "paiements": {},
            "mouvements_caisse": {"ecart": 0.0},
            "sequence": {"premier_ticket": "-", "dernier_ticket": "-"},
            "statut": {"caisse_juste": True}
        }
        lignes_z = [(f"{annee_str}-{mois_str}-01", json.dumps(virtual_z))]

        
    records = []
    totaux = {
        "ca_ttc": Decimal("0.00"),
        "htva_21": Decimal("0.00"),
        "tva_21": Decimal("0.00"),
        "htva_06": Decimal("0.00"),
        "tva_06": Decimal("0.00"),
        "bc_visa": Decimal("0.00"),
        "especes": Decimal("0.00"),
        "autres": Decimal("0.00"),
        "remb": Decimal("0.00"),
        "ecart": Decimal("0.00")
    }
    
    for date_z, json_str in lignes_z:
        data = json.loads(json_str)
        fin = data.get("financier", {})
        tva_dict = fin.get("tva", {})
        paiements = data.get("paiements", {})
        mouv = data.get("mouvements_caisse", {})
        
        t21 = tva_dict.get("0.21", {"base_htva": 0.0, "montant_tva": 0.0})
        t06 = tva_dict.get("0.06", {"base_htva": 0.0, "montant_tva": 0.0})
        
        c_ttc = Decimal(str(fin.get("ca_ttc", 0.0)))
        h21 = Decimal(str(t21["base_htva"]))
        tv21 = Decimal(str(t21["montant_tva"]))
        h06 = Decimal(str(t06["base_htva"]))
        tv06 = Decimal(str(t06["montant_tva"]))
        
        bc = Decimal(str(paiements.get("Bancontact/Visa", 0.0)))
        esp = Decimal(str(paiements.get("Espèces", 0.0)))
        aut = Decimal(str(paiements.get("Cartes Cadeaux / Autres", 0.0)))
        remb = Decimal(str(paiements.get("Remboursements", 0.0)))
        ec = Decimal(str(mouv.get("ecart", 0.0)))
        
        totaux["ca_ttc"] += c_ttc
        totaux["htva_21"] += h21
        totaux["tva_21"] += tv21
        totaux["htva_06"] += h06
        totaux["tva_06"] += tv06
        totaux["bc_visa"] += bc
        totaux["especes"] += esp
        totaux["autres"] += aut
        totaux["remb"] += remb
        totaux["ecart"] += ec
        
        seq = data.get("sequence", {})
        records.append({
            "Date": date_z,
            "Séquence Début": seq.get("premier_ticket", ""),
            "Séquence Fin": seq.get("dernier_ticket", ""),
            "CA Total TTC (€)": float(c_ttc),
            "Base HTVA 21% (€)": float(h21),
            "TVA 21% (€)": float(tv21),
            "Base HTVA 6% (€)": float(h06),
            "TVA 6% (€)": float(tv06),
            "Bancontact/Visa (€)": float(bc),
            "Espèces (€)": float(esp),
            "Autres Paiements (€)": float(aut),
            "Remboursements (€)": float(remb),
            "Écart de Caisse (€)": float(ec),
            "Statut": "Juste" if data.get("statut", {}).get("caisse_juste", False) else "Écart"
        })
        
    if format_type == "dict":
        return {
            "boutique": "L'ADRESSE B",
            "periode": f"{mois_str}/{annee_str}",
            "totaux": {k: float(v) for k, v in totaux.items()},
            "rapports_journaliers": records
        }
        
    if format_type == "json":
        res_obj = {
            "boutique": "L'ADRESSE B",
            "periode": f"{mois_str}/{annee_str}",
            "totaux": {k: float(v) for k, v in totaux.items()},
            "rapports_journaliers": records
        }
        _ensure_dir()
        path = os.path.join(EXPORT_DIR, f"export_comptable_{annee_str}_{mois_str}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(res_obj, f, ensure_ascii=False, indent=2)
        return path
        
    # Ligne de total
    total_row = {
        "Date": "TOTAL MENSUEL",
        "Séquence Début": "",
        "Séquence Fin": "",
        "CA Total TTC (€)": float(totaux["ca_ttc"]),
        "Base HTVA 21% (€)": float(totaux["htva_21"]),
        "TVA 21% (€)": float(totaux["tva_21"]),
        "Base HTVA 6% (€)": float(totaux["htva_06"]),
        "TVA 6% (€)": float(totaux["tva_06"]),
        "Bancontact/Visa (€)": float(totaux["bc_visa"]),
        "Espèces (€)": float(totaux["especes"]),
        "Autres Paiements (€)": float(totaux["autres"]),
        "Remboursements (€)": float(totaux["remb"]),
        "Écart de Caisse (€)": float(totaux["ecart"]),
        "Statut": ""
    }
    
    _ensure_dir()
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    headers = [
        "Date", "Séquence Début", "Séquence Fin", "CA Total TTC (€)", "Base HTVA 21% (€)",
        "TVA 21% (€)", "Base HTVA 6% (€)", "TVA 6% (€)", "Bancontact/Visa (€)", "Espèces (€)",
        "Autres Paiements (€)", "Remboursements (€)", "Écart de Caisse (€)", "Statut"
    ]
    
    if format_type == "csv":
        path = os.path.join(EXPORT_DIR, f"export_mensuel_{annee_str}_{mois_str}_{ts}.csv")
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers, delimiter=";")
            writer.writeheader()
            for r in records:
                r_copy = {k: str(v).replace(".", ",") if isinstance(v, float) else v for k, v in r.items()}
                writer.writerow(r_copy)
            t_copy = {k: str(v).replace(".", ",") if isinstance(v, float) else v for k, v in total_row.items()}
            writer.writerow(t_copy)
        return path
    else:
        path = os.path.join(EXPORT_DIR, f"export_mensuel_{annee_str}_{mois_str}_{ts}.xlsx")
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport Mensuel"
        ws.append(["RAPPORT COMPTABLE MENSUEL — L'ADRESSE B"])
        ws.append([f"Période : {mois_str}/{annee_str}"])
        ws.append([])
        ws.append(headers)
        for r in records:
            ws.append([r[h] for h in headers])
        ws.append([total_row[h] for h in headers])
        _style_header(ws, row=4, ncols=len(headers))
        wb.save(path)
        return path
